import tkinter as tk
import tkinter.messagebox
from datetime import datetime
from tkinter import *
from openpyxl import Workbook
from openpyxl.utils import get_column_letter,column_index_from_string 
from openpyxl import load_workbook
time=0
wb = Workbook()
ws = wb.active

filename = "myfile.xlsx"
wb = load_workbook(filename)
ws = wb.worksheets[0]

ws['A1'] = '流水號'
ws['B1'] = '姓名'
ws['C1'] = 'ID'
ws['D1'] = '時間'
ws['E1'] = '進出'
ws['F1'] = '總時間'
def callback(event):
    ID = ID_entry.get()
    count=0
    RPeople = open("people.txt", "r",encoding="utf-8")
    idtrue="f"
    for i in RPeople:
        tempid=i.split()
        if tempid[2]==ID:
            name=tempid[1]
            idtrue="true"
    if idtrue=="true": 
        f = open("rfid.txt", "r")
        inorout="  "
        for i in f:
            temp=i.split("\t")
            if temp[1]==ID:
                count=count+1
        if count%2==0:
            alltime=""
            WInplace = open("Inplace.txt", "a",encoding="utf-8")
            nowtime=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            WInplace.write(name+";;"+nowtime+";;"+'\n')
            WInplace.close()
            inorout=("  進館")
            time = tk.Label(app, text="           ",width=50,font=("Times", 15))
            time.place(x=60,y=180)
            RInplacetxt = open("Inplace.txt", "r",encoding="utf-8")
            text.delete(1.0, "end")
            for i in RInplacetxt:
                temp=i.split(";;")
                text.insert("insert",temp[0]+" "+temp[1]+"進場"+"\n", ("tag1"))
        if count%2==1:
            f1 = open("Inplace.txt", "r",encoding="utf-8")
            lines = f1.readlines()
            restart = open("Inplace.txt", "w",encoding="utf-8")
            for i in lines:
                temp=i.split(';;')
                if temp[0]==name:
                    d2=datetime.strptime(str(temp[1]),"%Y-%m-%d %H:%M:%S")
                    d1=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    d1=datetime.strptime(str(d1),"%Y-%m-%d %H:%M:%S")
                    timesecond = (d1 - d2).seconds
                    alltime=""
                    mins=0
                    seconds=0
                    if(timesecond<60):
                        time=timesecond
                        alltime=(str(time)+"秒")
                    elif(timesecond>60 and timesecond<=3600):
                        seconds=timesecond%60
                        time=int((timesecond/60))
                        alltime=(str(time)+"分"+str(seconds)+"秒")
                    elif(timesecond>3600):
                        mins=(timesecond/60)%60
                        seconds=(timesecond/3600)%60
                        time=int((timesecond/3600))
                        alltime=(str(time)+"小時"+str(mins)+"分"+str(seconds)+"秒")
                    lines.remove(temp[0]+";;"+temp[1]+";;"+'\n')
            inorout=("  出館")
            time = tk.Label(app, text="總時間"+alltime,borderwidth=1,bg="green2",font=("Times", 15))
            time.place(x=60,y=180)
            RInplacetxt = open("Inplace.txt", "r",encoding="utf-8")
            text.delete(1.0, "end")
            for i in lines:
                restart.write(i)
            restart.close()
            for i in RInplacetxt:
                temp=i.split(";;")
                text.insert("insert",temp[0]+" "+temp[1]+"進場"+"\n", ("tag1"))
        f = open("rfid.txt", "a")
        nowtime=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        f.write(name+"\t"+ID+'\t'+nowtime+inorout+'\n')
        ws.append([1,name,ID,nowtime,inorout,alltime])
        person = tk.Label(app, text=name+'\t'+inorout+'\t'+nowtime,borderwidth=1,bg="green2",font=("Times", 15))
        person.place(x = 60, y = 50)
        f.close()
    wb.save(filename)
    wb.close()
    ID_entry.delete(0, 'end')
    RInplacetxt = open("Inplace.txt", "r",encoding="utf-8")
    lines=0
    lines = len(RInplacetxt.readlines())
    people_label = tk.Label(app, text='現在人數  '+str(lines)+"   ",font=("Times", 15))
    people_label.place(x=320,y=110)
def center_window(root, width, height):  
    screenwidth = root.winfo_screenwidth()  
    screenheight = root.winfo_screenheight()  
    size = '%dx%d+%d+%d' % (width, height, (screenwidth - width)/2, (screenheight - height)/2)  
    root.geometry(size) 
app = tk.Tk()
Inplace = tk.Tk()
app.title('刷卡系統')
Inplace.title('場內')
center_window(app, 500, 240)
Inplace.geometry('600x600')

text = tk.Text(Inplace, width=200, height=200)
text.pack()
RInplacetxt = open("Inplace.txt", "r",encoding="utf-8")
lines = len(RInplacetxt.readlines())
RInplacetxt = open("Inplace.txt", "r",encoding="utf-8")
text.delete(1.0, "end")
text.tag_config("tag1", background="yellow", foreground="blue",font=("Times", 15))
for i in RInplacetxt:
    temp=i.split(";;")
    text.insert("insert",temp[0]+" "+temp[1]+"進場"+"\n", ("tag1"))
ID = tk.Frame(app)
ID.pack(side=tk.LEFT, ipadx=1, padx=10)
ID_label = tk.Label(ID, text='請刷卡',font=("Times", 15))
people_label = tk.Label(app, text='現在人數 '+str(lines),font=("Times", 15))
people_label.place(x=320,y=110)
ID_label.pack(side=tk.LEFT ,ipadx=10, padx=10)
ID_entry = tk.Entry(ID)
ID_entry.pack(side=tk.LEFT ,ipadx=10, padx=10)
app.bind('<Return>', callback)
app.mainloop()